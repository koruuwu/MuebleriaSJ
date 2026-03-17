# proyecto/admin_utils.py
from math import ceil
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter

from django.contrib import messages
from django.http import HttpResponseRedirect
from proyecto.utils.exception_logs import LogContext, write_exception_log

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

import os
import tempfile
import requests

from django.conf import settings

from reportlab.lib.utils import ImageReader
from reportlab.lib.units import cm

from openpyxl.drawing.image import Image as XLImage

from django import forms
from django.core.exceptions import ValidationError
from django.contrib import admin
from django.utils.html import format_html
from proyecto.supabase_client import subir_archivo

from django.contrib import admin
from django.contrib.admin.views.main import ChangeList
from django.urls import path
from django.http import HttpResponse
from django.utils import timezone

from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

class ExportReportMixin:
    

    export_report_name = "Reporte"
    export_filename_base = "reporte"
    export_rows_per_page = 35
    export_exclude_fields = ("vista_previa",)  # ejemplo, se puede override
    export_pdf_column_widths = None

    change_list_template = "admin/change_list_export.html"

        # ---------- Logo helpers ----------
    export_logo_url = "https://ckjwqojoaxxbttkfgtpt.supabase.co/storage/v1/object/public/logo/SanJoseLogo.png"         # puedes setearlo en settings o en cada admin
    export_logo_bucket_path = None # opcional si quieres construir URL desde supabase
    export_logo_height_cm = 1.8

    # =========================
    # PERMISOS DE EXPORTACIÓN
    # =========================
    def get_export_pdf_permission_codename(self):
        return f"export_pdf_{self.model._meta.model_name}"

    def get_export_excel_permission_codename(self):
        return f"export_excel_{self.model._meta.model_name}"

    def get_export_pdf_permission_full(self):
        return f"{self.model._meta.app_label}.{self.get_export_pdf_permission_codename()}"

    def get_export_excel_permission_full(self):
        return f"{self.model._meta.app_label}.{self.get_export_excel_permission_codename()}"

    def has_export_pdf_permission(self, request):
        return request.user.has_perm(self.get_export_pdf_permission_full())

    def has_export_excel_permission(self, request):
        return request.user.has_perm(self.get_export_excel_permission_full())

    def _redirect_no_export_permission(self, request, tipo_reporte):
        messages.error(
            request,
            f"No tienes permiso para exportar este reporte en formato {tipo_reporte}."
        )
        return HttpResponseRedirect(request.META.get("HTTP_REFERER", ".."))

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context["can_export_pdf"] = self.has_export_pdf_permission(request)
        extra_context["can_export_excel"] = self.has_export_excel_permission(request)
        return super().changelist_view(request, extra_context=extra_context)

    def _get_logo_url(self):
        """
        Prioridad:
        1) atributo en el admin/mixin: self.export_logo_url
        2) settings.EXPORT_REPORT_LOGO_URL
        """
        url = getattr(self, "export_logo_url", None)
        if url:
            return url
        return getattr(settings, "EXPORT_REPORT_LOGO_URL", None)

    def _download_logo_to_tmp(self, url: str) -> str | None:
        """
        Descarga el logo (URL pública) y lo cachea en un archivo temporal.
        Devuelve path local o None si falla.
        """
        try:
            # cache simple: mismo nombre por URL (hash)
            key = str(abs(hash(url)))
            cache_dir = os.path.join(tempfile.gettempdir(), "muebleria_report_cache")
            os.makedirs(cache_dir, exist_ok=True)
            local_path = os.path.join(cache_dir, f"logo_{key}.png")

            if os.path.exists(local_path) and os.path.getsize(local_path) > 0:
                return local_path

            r = requests.get(url, timeout=10)
            r.raise_for_status()
            with open(local_path, "wb") as f:
                f.write(r.content)
            return local_path
        except Exception:
            return None

    def _get_logo_local_path(self) -> str | None:
        url = self._get_logo_url()
        if not url:
            return None
        return self._download_logo_to_tmp(url)

    def get_urls(self):
        """
        Añade las URLs /export/pdf/ y /export/excel/ para este ModelAdmin.
        """
        urls = super().get_urls()
        # app_label y model_name para no repetir nombres
        app_label = self.model._meta.app_label
        model_name = self.model._meta.model_name

        custom = [
            path(
                "export/pdf/",
                self.admin_site.admin_view(self.export_pdf),
                name=f"{app_label}_{model_name}_export_pdf",
            ),
            path(
                "export/excel/",
                self.admin_site.admin_view(self.export_excel),
                name=f"{app_label}_{model_name}_export_excel",
            ),
        ]
        return custom + urls

    # ---------- Helpers comunes ----------
    def _get_changelist_queryset(self, request):
        """
        Devuelve el queryset tal como aparece en el changelist:
        respeta búsqueda, filtros, orden, etc.
        """
        cl = self.get_changelist_instance(request)
        return cl.get_queryset(request)
    
    def _get_export_columns(self):
        """
        Devuelve la lista de 'column keys' que usaremos para el reporte.
        Prioridad:
        1) self.export_fields (si existe)
        2) self.list_display
        3) todos los campos del modelo
        Aplicando export_exclude_fields.
        """
        # 1) export_fields explícitos en el admin (si los defines)
        if hasattr(self, "export_fields") and self.export_fields:
            cols = list(self.export_fields)
        # 2) list_display del admin
        elif getattr(self, "list_display", None):
            cols = [c for c in self.list_display if c != "action_checkbox"]
        # 3) todos los campos del modelo
        else:
            cols = [f.name for f in self.model._meta.fields]

        excl = getattr(self, "export_exclude_fields", ())
        cols = [c for c in cols if c not in excl]
        return cols

    def _export_get_value(self, obj, field_name):
        """
        Obtiene el valor de una 'columna' para un objeto:
        - Si el admin tiene un método con ese nombre -> lo llama con obj.
        - Si el modelo tiene atributo/propiedad -> lo usa (si es callable, lo llama).
        """
        # método definido en el ModelAdmin (por ejemplo total_pedidos, vista_previa, etc.)
        if hasattr(self, field_name):
            admin_attr = getattr(self, field_name)
            if callable(admin_attr):
                return admin_attr(obj)

        # atributo / propiedad del modelo
        attr = getattr(obj, field_name, "")
        if callable(attr):
            return attr()
        return attr

    def _wrap_text(self, text, max_width, font_name="Helvetica", font_size=10):
        """
        Divide 'text' en varias líneas para que cada línea no sobrepase max_width.
        Devuelve una lista de líneas.
        """
        if text is None:
            return [""]

        text = str(text)
        words = text.split()
        if not words:
            return [""]

        lines = []
        current = ""

        for word in words:
            candidate = word if not current else current + " " + word
            if stringWidth(candidate, font_name, font_size) <= max_width:
                current = candidate
            else:
                if current:
                    lines.append(current)
                # Si la palabra sola ya es muy ancha, la partimos “a lo bruto”.
                if stringWidth(word, font_name, font_size) <= max_width:
                    current = word
                else:
                    fragment = ""
                    for ch in word:
                        cand = fragment + ch
                        if stringWidth(cand, font_name, font_size) <= max_width:
                            fragment = cand
                        else:
                            if fragment:
                                lines.append(fragment)
                            fragment = ch
                    current = fragment

        if current:
            lines.append(current)

        return lines or [""]

    def get_export_queryset(self, request):
        # Usa el mismo queryset que ve el usuario en el changelist
        cl = self.get_changelist_instance(request)
        return cl.get_queryset(request)

    def get_export_columns(self, request):
        """
        Devuelve la lista de nombres de campos / métodos a exportar.
        Por defecto: list_display sin 'action_checkbox' ni los de export_exclude_fields.
        Si el admin define export_columns, se usa eso.
        """
        if hasattr(self, "export_columns"):
            return list(self.export_columns)

        cols = []
        for name in self.get_list_display(request):
            if name == "action_checkbox":
                continue
            if name in getattr(self, "export_exclude_fields", ()):
                continue
            cols.append(name)
        return cols

    def get_export_headers(self, request, columns):
        headers = []

        for name in columns:
            try:
                # Si es campo directo del modelo
                field = self.model._meta.get_field(name)
                headers.append(str(field.verbose_name).capitalize())
                continue
            except Exception:
                pass

            # Si es FK tipo id_mueble pero quieres usar el campo real
            if name.startswith("id_"):
                real_name = name.replace("id_", "")
                try:
                    field = self.model._meta.get_field(name)
                    headers.append(str(field.verbose_name).capitalize())
                    continue
                except Exception:
                    pass

            # Si es método con short_description
            attr = getattr(self, name, None) or getattr(self.model, name, None)
            if attr and hasattr(attr, "short_description"):
                headers.append(str(attr.short_description))
            else:
                headers.append(name.replace("_", " ").capitalize())

        return headers

    def get_export_row(self, obj, columns):
        """
        Devuelve una lista con los valores de cada columna para un registro.
        """
        data = []
        for name in columns:
            value = getattr(obj, name, "")
            if callable(value):
                try:
                    value = value()
                except TypeError:
                    # por si requiere argumentos, ignoramos
                    pass
            if value is None:
                value = ""
            data.append(str(value))
        return data

    # ---------- PDF ----------

        
    def export_pdf(self, request):
        if not self.has_export_pdf_permission(request):
            return self._redirect_no_export_permission(request, "PDF")

        try:
            qs = self._get_changelist_queryset(request)

            

            response = HttpResponse(content_type="application/pdf")
            response["Content-Disposition"] = f'attachment; filename="{self.export_filename_base}.pdf"'

            c = canvas.Canvas(response, pagesize=letter)
            page_width, page_height = letter

            empresa = "Mueblería San José"
            reporte = getattr(self, "export_report_name", "Reporte")
            usuario = request.user.get_username()
            ahora = timezone.localtime()
            logo_path = self._get_logo_local_path()

            # Márgenes y layout
            left_margin = 40
            right_margin = 40
            bottom_margin = 60
            
            top_header_y = page_height - 50
            table_top_y = page_height - 110
            bottom_limit_y = bottom_margin + 20
            line_height = 14

            total_width = page_width - left_margin - right_margin

            # ----- Columnas -----
            columns = self._get_export_columns()
            headers = self.get_export_headers(request, columns)
            num_cols = len(columns)

            if self.export_pdf_column_widths and len(self.export_pdf_column_widths) == num_cols:
                col_widths = self.export_pdf_column_widths
            else:
                col_width = total_width / max(num_cols, 1)
                col_widths = [col_width] * num_cols

            col_x = [left_margin]
            for w_col in col_widths[:-1]:
                col_x.append(col_x[-1] + w_col)

            # ----- Wrap de encabezados y altura dinámica del header -----
            header_lines_per_col = []
            for idx, titulo in enumerate(headers):
                lines = self._wrap_text(
                    str(titulo),
                    max_width=col_widths[idx] - 4,
                    font_name="Helvetica-Bold",
                    font_size=10,
                )
                header_lines_per_col.append(lines)

            max_header_lines = max((len(lines) for lines in header_lines_per_col), default=1)
            table_header_height = max_header_lines * line_height

            # Inicio real de datos debajo del encabezado
            header_bottom_y = table_top_y - table_header_height - 2
            data_top_y = header_bottom_y - 8

            # ----- Encabezado y pie -----
            def header():
                text_x = left_margin

                # Texto primero (lado izquierdo)
                c.setFont("Helvetica-Bold", 14)
                c.drawString(text_x, top_header_y, empresa)

                c.setFont("Helvetica", 12)
                c.drawString(text_x, top_header_y - 20, reporte)

                # Logo a la derecha
                if logo_path:
                    try:
                        img = ImageReader(logo_path)
                        logo_h = self.export_logo_height_cm * cm
                        logo_w = logo_h  # cuadrado, ajusta proporción

                        logo_x = page_width - right_margin - logo_w
                        logo_y = top_header_y - logo_h + 5

                        c.drawImage(
                            img,
                            logo_x,
                            logo_y,
                            width=logo_w,
                            height=logo_h,
                            preserveAspectRatio=True,
                            mask="auto",
                        )
                    except Exception:
                        pass

                # Línea divisoria debajo de todo el encabezado de columnas
                header_bottom_y = table_top_y - table_header_height - 2
                c.line(left_margin, header_bottom_y, page_width - right_margin, header_bottom_y)

            def footer(page_num, total_pages):
                c.line(left_margin, 60, page_width - right_margin, 60)
                c.setFont("Helvetica", 9)

                texto_pagina = f"Página {page_num} de {total_pages or 1}"
                texto_fecha = f"Fecha y hora: {ahora.strftime('%Y-%m-%d %H:%M:%S')}"
                texto_usuario = f"Usuario: {usuario}"

                c.drawString(left_margin, 45, texto_usuario)                  # izquierda
                c.drawCentredString(page_width / 2, 45, texto_pagina)         # centro ✅
                c.drawRightString(page_width - right_margin, 45, texto_fecha) # derecha

            # Y de inicio de datos (parte superior de la primera fila)
            

            # ---------- 1) SIMULACIÓN para saber cuántas páginas habrá ----------
            page_count = 1
            y_sim = data_top_y

            for obj in qs:
                max_lines = 1
                for idx, field_name in enumerate(columns):
                    value = self._export_get_value(obj, field_name)
                    lines = self._wrap_text(
                        value,
                        max_width=col_widths[idx] - 4,
                        font_name="Helvetica",
                        font_size=10,
                    )
                    max_lines = max(max_lines, len(lines))

                row_height = max_lines * line_height
                if y_sim - row_height < bottom_limit_y:
                    page_count += 1
                    y_sim = data_top_y

                y_sim -= row_height

            total_pages = max(page_count, 1)

            # ---------- 2) DIBUJO real ----------
            page_num = 1
            header()

            # Encabezado de columnas
            c.setFont("Helvetica-Bold", 10)
            header_y = table_top_y

            for i, lines in enumerate(header_lines_per_col):
                for line_idx, line in enumerate(lines):
                    c.drawString(
                        col_x[i],
                        header_y - (line_idx * line_height),
                        line,
                    )

            c.setFont("Helvetica", 10)

            y = data_top_y  # parte superior de la primera fila
            row_index = 0   # para zebra

            for obj in qs:
                # Valores de la fila según columnas
                values = self.get_export_row(obj, columns)

                # Envolver texto por columna
                col_lines = []
                max_lines_in_row = 1
                for value, max_w in zip(values, col_widths):
                    text = "" if value is None else str(value)
                    lines = self._wrap_text(text, max_w - 4)
                    if not lines:
                        lines = [""]
                    col_lines.append(lines)
                    max_lines_in_row = max(max_lines_in_row, len(lines))

                row_height = max_lines_in_row * line_height

                # ¿Cabe la fila completa? Si no, salto de página
                if y - row_height < bottom_limit_y:
                    footer(page_num, total_pages)
                    c.showPage()
                    page_num += 1
                    header()

                    # reimprimir encabezado de columnas en la nueva página
                    c.setFont("Helvetica-Bold", 10)
                    header_y = table_top_y

                    for i, lines in enumerate(header_lines_per_col):
                        for line_idx, line in enumerate(lines):
                            c.drawString(
                                col_x[i],
                                header_y - (line_idx * line_height),
                                line,
                            )

                    c.setFont("Helvetica", 10)
                    y = data_top_y  # reset parte superior de filas

                row_top = y
                row_bottom = y - row_height

                # Fondo “zebra” para la fila completa
                if row_index % 2 == 0:
                    c.setFillColor(colors.whitesmoke)
                    c.rect(
                        left_margin,
                        row_bottom,
                        total_width,
                        row_height,
                        stroke=0,
                        fill=1,
                    )
                    c.setFillColor(colors.black)

                # Texto de la fila, línea por línea
                for line_idx in range(max_lines_in_row):
                    text_y = row_top - (line_idx + 1) * line_height + 3  # +3 = pequeño padding
                    for col_idx, lines in enumerate(col_lines):
                        if line_idx < len(lines):
                            c.drawString(
                                col_x[col_idx],
                                text_y,
                                lines[line_idx],
                            )

                y = row_bottom
                row_index += 1

            footer(page_num, total_pages)
            c.save()
            return response
        except Exception as exc:
            write_exception_log(
                LogContext(
                    modulo=self.model._meta.app_label,
                    request=request,
                ),
                exc
            )
            messages.error(request, "Ocurrió un error al exportar el PDF.")
            return HttpResponseRedirect(request.META.get("HTTP_REFERER", ".."))

    # ---------- Export Excel ----------

    def export_excel(self, request):
        if not self.has_export_excel_permission(request):
            return self._redirect_no_export_permission(request, "Excel")

        try:
            qs = self._get_changelist_queryset(request)
            columns = self._get_export_columns()
            headers = self.get_export_headers(request, columns)

            wb = Workbook()
            ws = wb.active
            ws.title = str(self.export_report_name)[:31]

            empresa = getattr(self, "export_company_name", "Mueblería San José")
            reporte = getattr(self, "export_report_name", "Reporte")
            usuario = request.user.get_username()
            ahora = timezone.localtime().strftime("%Y-%m-%d %H:%M:%S")

            # ---------- Estilos ----------
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.drawing.image import Image as XLImage

            bold_16 = Font(bold=True, size=16)
            bold_12 = Font(bold=True, size=12)
            bold_11 = Font(bold=True, size=11)
            normal_10 = Font(size=10)

            align_left = Alignment(horizontal="left", vertical="center")
            align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            align_wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

            fill_header = PatternFill("solid", fgColor="E9ECEF")   # gris suave
            fill_zebra = PatternFill("solid", fgColor="F7F7F7")    # zebra gris más claro

            thin = Side(style="thin", color="BFBFBF")
            border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

            # ---------- Logo ----------
            logo_path = self._get_logo_local_path()  # usa el mismo helper
            # Layout: logo en A1, texto en B1...
            text_col = "B"

            # Espacio de header (filas 1-5)
            ws.row_dimensions[1].height = 45
            ws.row_dimensions[2].height = 22
            ws.row_dimensions[3].height = 18
            ws.row_dimensions[4].height = 18
            ws.row_dimensions[5].height = 10

            if logo_path:
                try:
                    from PIL import Image as PILImage

                    # Abrimos imagen para conocer tamaño real
                    pil_img = PILImage.open(logo_path)
                    orig_width, orig_height = pil_img.size

                    img = XLImage(logo_path)

                    # Altura deseada en px
                    desired_height = 80

                    # Mantener proporción
                    ratio = orig_width / orig_height
                    img.height = desired_height
                    img.width = desired_height * ratio

                    ws.add_image(img, "A1")

                    # Ajustar ancho columna A según logo
                    ws.column_dimensions["A"].width = 14

                except Exception:
                    ws.column_dimensions["A"].width = 4

            # ---------- Header texto ----------
            ws[f"{text_col}1"] = empresa
            ws[f"{text_col}1"].font = bold_16
            ws[f"{text_col}1"].alignment = align_left

            ws[f"{text_col}2"] = reporte
            ws[f"{text_col}2"].font = bold_12
            ws[f"{text_col}2"].alignment = align_left

            ws[f"{text_col}3"] = f"Usuario: {usuario}"
            ws[f"{text_col}3"].font = normal_10
            ws[f"{text_col}3"].alignment = align_left

            ws[f"{text_col}4"] = f"Fecha y hora: {ahora}"
            ws[f"{text_col}4"].font = normal_10
            ws[f"{text_col}4"].alignment = align_left

            # Línea visual separadora (fila 5)
            max_cols = len(headers)
            if max_cols < 1:
                max_cols = 1
            # une B5 hasta última columna del reporte para "subrayar"
            from openpyxl.utils import get_column_letter
            last_col_letter = get_column_letter(max_cols + 1)  # +1 porque empezamos tabla en col 1, pero header texto en B
            ws.merge_cells(f"{text_col}5:{last_col_letter}5")
            ws[f"{text_col}5"] = ""
            ws[f"{text_col}5"].border = Border(bottom=thin)

            # ---------- Tabla ----------
            start_row = 6
            start_col = 1  # A

            # Encabezados de la tabla
            for idx, h in enumerate(headers, start=start_col):
                cell = ws.cell(row=start_row, column=idx, value=h)
                cell.font = bold_11
                cell.alignment = align_center
                cell.fill = fill_header
                cell.border = border_thin

            # Filas de datos
            row = start_row + 1
            for row_idx, obj in enumerate(qs):
                values = self.get_export_row(obj, columns)

                zebra = (row_idx % 2 == 0)  # intercalado como en PDF (0 gris, 1 blanco)
                for col_idx, val in enumerate(values, start=start_col):
                    cell = ws.cell(row=row, column=col_idx, value=val)
                    cell.font = normal_10
                    cell.alignment = align_wrap_left
                    cell.border = border_thin
                    if zebra:
                        cell.fill = fill_zebra
                row += 1

            end_row = row - 1
            end_col = start_col + len(headers) - 1

            # Ajuste de anchos por contenido (con límites para que no se dispare)
            max_width = 35
            min_width = 12

            for col_idx in range(start_col, start_col + len(headers)):
                col_letter = get_column_letter(col_idx)
                max_len = 0

                # medir header + celdas (limitado)
                header_val = ws.cell(row=start_row, column=col_idx).value
                if header_val:
                    max_len = max(max_len, len(str(header_val)))

                for r in range(start_row + 1, end_row + 1):
                    v = ws.cell(row=r, column=col_idx).value
                    if v is None:
                        continue
                    max_len = max(max_len, len(str(v)))

                # factor simple para que se vea bien
                width = max(min_width, min(max_width, max_len + 2))
                ws.column_dimensions[col_letter].width = width

            # Congelar encabezado de tabla (para que siempre se vea)
            ws.freeze_panes = ws["A7"]  # fila después de headers

            # ---------- Respuesta ----------
            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            filename_base = getattr(self, "export_filename_base", "reporte")
            filename = f"{filename_base}_reporte.xlsx"
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

                    # ---------------------------
            # Formato de tabla (bordes + encabezado + zebra)
            # ---------------------------
            thin = Side(style="thin")
            border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

            header_fill = PatternFill("solid", fgColor="D9D9D9")   # gris claro
            zebra_fill = PatternFill("solid", fgColor="F2F2F2")    # gris más suave

            header_font = Font(bold=True)
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_align = Alignment(vertical="top", wrap_text=True)

            # Rango de tabla
            last_row = row - 1                      # 'row' ya quedó apuntando a la siguiente fila libre
            first_col = 1
            last_col = len(headers)

            # Encabezados (fila start_row)
                    # Encabezados (fila start_row)
            for col in range(first_col, last_col + 1):
                cell = ws.cell(row=start_row, column=col)

                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

                # Forzar borde inferior visible
                cell.border = Border(
                    left=thin,
                    right=thin,
                    top=thin,
                    bottom=Side(style="medium")  # más grueso para que destaque
                )

            # Datos (desde start_row+1 hasta last_row)
            for r in range(start_row + 1, last_row + 1):
                # zebra: primera fila de datos gris, siguiente blanco, etc.
                use_zebra = ((r - (start_row + 1)) % 2 == 0)

                for ccol in range(first_col, last_col + 1):
                    cell = ws.cell(row=r, column=ccol)
                    cell.border = border_all
                    cell.alignment = cell_align
                    if use_zebra:
                        cell.fill = zebra_fill

            # Ajuste de altura del header para que se vea mejor
            ws.row_dimensions[start_row].height = 20

            wb.save(response)
            return response
        except Exception as exc:
            write_exception_log(
                LogContext(
                    modulo=self.model._meta.app_label,
                    request=request,
                ),
                exc
            )
            messages.error(request, "Ocurrió un error al exportar el Excel.")
            return HttpResponseRedirect(request.META.get("HTTP_REFERER", ".."))

class PaginacionAdminMixin:

    actions = ["set_pagination_1", "set_pagination_10", "set_pagination_25", "set_pagination_50", "set_pagination_100"]

    def set_pagination_1(self, request, queryset):
        request.session['per_page'] = 1
        self.message_user(request, "Mostrando 1 elementos por página.")
    set_pagination_1.short_description = "Mostrar 1 por página"

    def set_pagination_10(self, request, queryset):
        request.session['per_page'] = 10
        self.message_user(request, "Mostrando 10 elementos por página.")
    set_pagination_10.short_description = "Mostrar 10 por página"

    def set_pagination_25(self, request, queryset):
        request.session['per_page'] = 25
        self.message_user(request, "Mostrando 25 elementos por página.")
    set_pagination_25.short_description = "Mostrar 25 por página"

    def set_pagination_50(self, request, queryset):
        request.session['per_page'] = 50
        self.message_user(request, "Mostrando 50 elementos por página.")
    set_pagination_50.short_description = "Mostrar 50 por página"

    def set_pagination_100(self, request, queryset):
        request.session['per_page'] = 100
        self.message_user(request, "Mostrando 100 elementos por página.")
    set_pagination_100.short_description = "Mostrar 100 por página"

    def changelist_view(self, request, extra_context=None):
        if 'per_page' in request.session:
            self.list_per_page = request.session['per_page']
        else:
            self.list_per_page = 10
        return super().changelist_view(request, extra_context=extra_context)





class AdminConImagenMixin:
    readonly_fields = ("vista_previa", "imagen_url", "imagen")
    bucket_name = "default"  # Cada admin puede sobrescribir este valor

    def save_model(self, request, obj, form, change):
        archivo = form.cleaned_data.get("archivo_temp")
        if archivo:
            ruta, url = subir_archivo(archivo, self.bucket_name)
            obj.imagen = ruta
            obj.imagen_url = url
        super().save_model(request, obj, form, change)

    def vista_previa(self, obj):
        if getattr(obj, "imagen_url", None):
            return format_html(
                '<img src="{}" style="max-height:100px;border-radius:8px;" />',
                obj.imagen_url,
            )
        return "Sin imagen"
    vista_previa.short_description = "Vista previa"



from django.contrib import admin
from django.core.exceptions import ValidationError

class UniqueFieldAdminMixin:
    """
    Mixin para ModelAdmin que valida campos únicos antes de guardar,
    sin necesidad de definir un ModelForm.
    """
    
    # Campo único o lista de campos únicos a validar
    unique_fields = []  # ['tipo_documento'] o ['campo1', 'campo2']

    def get_form(self, request, obj=None, **kwargs):
        """
        Inyecta dinámicamente la validación de duplicados en el form que genera el admin.
        """
        form = super().get_form(request, obj, **kwargs)

        unique_fields = self.unique_fields  # lista de campos a validar

        class _FormWithUniqueValidation(form):
            def clean(self_inner):
                cleaned_data = super(_FormWithUniqueValidation, self_inner).clean()
                for field in unique_fields:
                    valor = cleaned_data.get(field)
                    if valor is None:
                        continue
                    qs = self_inner._meta.model.objects.exclude(pk=self_inner.instance.pk)
                    if qs.filter(**{field: valor}).exists():
                        raise ValidationError({field: f'⚠️ El valor "{valor}" ya existe.'})
                return cleaned_data

        return _FormWithUniqueValidation


